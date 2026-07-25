"""WheelAthlete export-schema-2 processor for paired IMU XLSX and C3D data.

The source folders are read-only inputs. The processor repairs malformed XLSX
dimensions, reconstructs an arrival-time timeline without using the invalid
``timestamp_utc_ms`` field, writes exact raw/training CSV contracts, and emits
alignment/QC provenance for every trial.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import sys
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
from openpyxl import load_workbook

try:
    import ezc3d
except ImportError:  # pragma: no cover - exercised by CLI dependency check
    ezc3d = None


PROCESSOR_VERSION = "2.0.0"
RAW_HEADER = ["time_us", "ax_g", "ay_g", "az_g", "gx_dps", "gy_dps", "gz_dps"]
TRAINING_HEADER = [
    "time_us",
    "left_ax_g", "left_ay_g", "left_az_g",
    "left_gx_dps", "left_gy_dps", "left_gz_dps",
    "right_ax_g", "right_ay_g", "right_az_g",
    "right_gx_dps", "right_gy_dps", "right_gz_dps",
]

C3D_BY_TRIAL = {
    ("slope", 1): "10Slope 1.c3d",
    ("10x5", 1): "10x5m 1.c3d",
    ("clock_wise", 1): "360CR 1.c3d",
    ("anti_clock_wise", 1): "360AR 1.c3d",
    ("sprint_5m", 1): "5MST.c3d",
    ("one_stroke", 2): "One ST2.c3d",
    ("slalom", 1): "Sl.c3d",
    ("slalom", 2): "Sl 2.c3d",
}


@dataclass(frozen=True)
class WheelData:
    side: str
    seq: np.ndarray
    arrival_ms: np.ndarray
    time_us: np.ndarray
    values: np.ndarray
    gaps: int
    fit_slope_ms: float
    fit_rms_ms: float


@dataclass(frozen=True)
class Alignment:
    lag_frames: int
    scale: float
    drift_ppm: float
    correlation: float
    peak_margin: float
    confidence: str
    repeatability_frames: int


def normalize_label(value: str) -> str:
    label = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    aliases = {"slalop": "slalom", "anti_clockwise": "anti_clock_wise", "clockwise": "clock_wise"}
    return aliases.get(label, label)


def parse_trial_name(path: Path) -> tuple[str, int]:
    stem = path.stem
    match = re.search(r"(?:-|_)trial_(\d+)$", stem, flags=re.IGNORECASE)
    if not match:
        raise ValueError(f"Cannot parse trial number from {path.name}")
    trial = int(match.group(1))
    prefix = stem[: match.start()]
    prefix = re.sub(r"(?:_|-)10-7-2026$", "", prefix, flags=re.IGNORECASE)
    return normalize_label(prefix), trial


def unwrap_uint32(values: Sequence[int]) -> np.ndarray:
    result = np.empty(len(values), dtype=np.int64)
    if not values:
        return result
    offset = 0
    previous = int(values[0])
    result[0] = previous
    for index, raw_value in enumerate(values[1:], 1):
        value = int(raw_value)
        if value < previous and previous - value > 0x80000000:
            offset += 0x100000000
        result[index] = value + offset
        previous = value
    return result


def robust_sequence_fit(seq: np.ndarray, arrival_ms: np.ndarray) -> tuple[float, float, float]:
    """Robust ``arrival_ms = intercept + slope*seq`` fit.

    Phone arrival times are batchy and may repeat. Centered iterative fitting
    prevents those duplicates and isolated scheduling outliers from changing
    the reconstructed sample interval.
    """
    if len(seq) < 2:
        return float(arrival_ms[0] if len(arrival_ms) else 0), 10.0, 0.0
    x = seq.astype(np.float64)
    y = arrival_ms.astype(np.float64)
    x0 = float(np.median(x))
    y0 = float(np.median(y))
    mask = np.isfinite(x) & np.isfinite(y)
    slope = 10.0
    intercept = y0 - slope * x0
    for _ in range(5):
        xc = x[mask] - x0
        yc = y[mask] - y0
        denom = float(np.dot(xc, xc))
        if denom > 0:
            candidate = float(np.dot(xc, yc) / denom)
            if 2.0 <= candidate <= 25.0:
                slope = candidate
        intercept = float(np.median(y[mask] - slope * x[mask]))
        residual = y - (intercept + slope * x)
        median = float(np.median(residual[mask]))
        mad = float(np.median(np.abs(residual[mask] - median)))
        threshold = max(2.5, 4.5 * 1.4826 * mad)
        new_mask = np.isfinite(residual) & (np.abs(residual - median) <= threshold)
        if new_mask.sum() < max(2, len(x) // 2) or np.array_equal(mask, new_mask):
            break
        mask = new_mask
    residual = y[mask] - (intercept + slope * x[mask])
    rms = float(np.sqrt(np.mean(residual * residual))) if residual.size else 0.0
    return intercept, slope, rms


def _numeric(value: object) -> float:
    number = float(value)
    if not math.isfinite(number):
        raise ValueError("non-finite value")
    return number


def read_legacy_xlsx(path: Path) -> dict[str, WheelData]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    provisional: dict[str, tuple[np.ndarray, np.ndarray, np.ndarray, int, float, float, float]] = {}
    for worksheet in workbook.worksheets:
        # July exports incorrectly declare <dimension ref="A1"/> despite
        # containing thousands of rows.
        worksheet.reset_dimensions()
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header = [str(value).strip() if value is not None else "" for value in next(iterator)]
        except StopIteration:
            continue
        index = {name: position for position, name in enumerate(header)}
        required = {"seq", "timestamp_app_ms", "ax", "ay", "az", "gx", "gy", "gz"}
        if not required.issubset(index):
            continue
        seq_values: list[int] = []
        arrivals: list[float] = []
        values: list[list[float]] = []
        side = worksheet.title.strip().upper()[:1]
        for row in iterator:
            try:
                seq_values.append(int(row[index["seq"]]))
                arrivals.append(_numeric(row[index["timestamp_app_ms"]]))
                values.append([_numeric(row[index[name]]) for name in ("ax", "ay", "az", "gx", "gy", "gz")])
                if "wheel" in index and row[index["wheel"]] is not None:
                    side = str(row[index["wheel"]]).strip().upper()[:1]
            except (IndexError, TypeError, ValueError):
                continue
        if side not in {"L", "R"} or not seq_values:
            continue
        seq = unwrap_uint32(seq_values)
        arrival = np.asarray(arrivals, dtype=np.float64)
        axes = np.asarray(values, dtype=np.float64)
        intercept, slope, rms = robust_sequence_fit(seq, arrival)
        predicted_first = intercept + slope * float(seq[0])
        gaps = int(np.maximum(np.diff(seq) - 1, 0).sum()) if len(seq) > 1 else 0
        provisional[side] = (seq, arrival, axes, gaps, intercept, slope, rms, predicted_first)
    workbook.close()
    if not provisional:
        raise ValueError(f"No L/R sample worksheets found in {path}")
    common_start_ms = min(item[7] for item in provisional.values())
    result: dict[str, WheelData] = {}
    for side, item in provisional.items():
        seq, arrival, axes, gaps, intercept, slope, rms, _ = item
        time_us = np.rint((intercept + slope * seq - common_start_ms) * 1000.0).astype(np.int64)
        result[side] = WheelData(side, seq, arrival, time_us, axes, gaps, slope, rms)
    return result


def _robust_normalize(values: np.ndarray) -> np.ndarray:
    finite = values[np.isfinite(values)]
    if finite.size == 0:
        return np.zeros_like(values, dtype=np.float64)
    median = float(np.median(finite))
    mad = float(np.median(np.abs(finite - median)))
    scale = max(1e-9, 1.4826 * mad)
    normalized = (np.nan_to_num(values, nan=median) - median) / scale
    return np.clip(normalized, -8.0, 8.0)


def _smooth(values: np.ndarray, window: int) -> np.ndarray:
    if window <= 1 or len(values) < window:
        return values.astype(np.float64, copy=True)
    kernel = np.ones(window, dtype=np.float64) / window
    return np.convolve(values, kernel, mode="same")


def imu_envelope(times_us: np.ndarray, values: np.ndarray, rate_hz: int = 100) -> np.ndarray:
    accel = np.linalg.norm(values[:, :3], axis=1)
    gyro = np.linalg.norm(values[:, 3:], axis=1)
    motion = np.abs(_robust_normalize(accel)) + 0.35 * np.abs(_robust_normalize(gyro))
    return _smooth(motion, max(3, rate_hz // 5))


def read_c3d_envelope(path: Path) -> tuple[np.ndarray, float, dict[str, object]]:
    if ezc3d is None:
        raise RuntimeError("ezc3d is required: pip install ezc3d")
    c3d = ezc3d.c3d(str(path))
    point_rate = float(c3d["parameters"]["POINT"]["RATE"]["value"][0])
    points = np.asarray(c3d["data"]["points"], dtype=np.float64)
    xyz = np.moveaxis(points[:3], 2, 0)  # frames, xyz, markers
    valid = np.isfinite(xyz).all(axis=1) & (np.linalg.norm(xyz, axis=1) > 1e-6)
    velocity = np.linalg.norm(np.diff(xyz, axis=0), axis=1) * point_rate
    velocity_valid = valid[1:] & valid[:-1]
    velocity[~velocity_valid] = np.nan
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)
        marker_speed = np.nanmedian(velocity, axis=1)
    marker_speed = np.r_[marker_speed[0] if marker_speed.size else 0.0, marker_speed]
    marker_coverage = float(np.mean(valid)) if valid.size else 0.0

    analog = np.asarray(c3d["data"]["analogs"], dtype=np.float64)
    analog_labels = [str(value).strip() for value in c3d["parameters"]["ANALOG"]["LABELS"]["value"]]
    force_indices = [
        index for index, label in enumerate(analog_labels)
        if re.match(r"^(F[xyz]|Force)", label, flags=re.IGNORECASE)
    ]
    force_envelope = np.zeros(points.shape[2], dtype=np.float64)
    if analog.size and force_indices:
        # ezc3d analog layout is subframes x channels x point-frames.
        if analog.ndim == 3:
            selected = analog[:, force_indices, :]
            force = np.sqrt(np.sum(selected * selected, axis=1)).mean(axis=0)
        else:
            flat = analog.reshape(analog.shape[-2], -1)
            force = np.sqrt(np.sum(flat[force_indices] ** 2, axis=0))
            source_x = np.linspace(0.0, 1.0, len(force), endpoint=True)
            force = np.interp(np.linspace(0.0, 1.0, len(force_envelope)), source_x, force)
        if len(force) == len(force_envelope):
            force_envelope = force

    marker_component = np.abs(_robust_normalize(marker_speed))
    force_component = np.abs(_robust_normalize(force_envelope)) if np.any(force_envelope) else 0.0
    envelope = _smooth(marker_component + 0.5 * force_component, max(3, round(point_rate / 5)))
    return envelope, point_rate, {
        "point_rate_hz": point_rate,
        "frames": int(len(envelope)),
        "marker_coverage": marker_coverage,
        "force_channels": [analog_labels[index] for index in force_indices],
    }


def detect_three_pulse_cue(envelope: np.ndarray, rate_hz: int = 100) -> int | None:
    """Detect three separated motion pulses bracketed by quiet periods."""
    search = envelope[: min(len(envelope), rate_hz * 10)]
    if len(search) < rate_hz * 3:
        return None
    baseline = float(np.median(search))
    mad = float(np.median(np.abs(search - baseline)))
    threshold = baseline + max(0.8, 3.0 * 1.4826 * mad)
    candidates: list[int] = []
    above = np.flatnonzero(search >= threshold)
    if above.size:
        starts = np.r_[above[0], above[1:][np.diff(above) > 1]]
        ends = np.r_[above[:-1][np.diff(above) > 1], above[-1]]
        for start, end in zip(starts, ends):
            segment = search[start : end + 1]
            peak = float(segment.max())
            plateau = np.flatnonzero(segment == peak)
            candidates.append(int(start + round(float(plateau.mean()))))
    for start in range(max(0, len(candidates) - 2)):
        group = candidates[start : start + 3]
        if len(group) < 3 or group[-1] - group[0] > rate_hz * 5:
            continue
        before = search[max(0, group[0] - rate_hz) : max(0, group[0] - rate_hz // 4)]
        after = search[min(len(search), group[-1] + rate_hz // 4) : min(len(search), group[-1] + rate_hz)]
        if before.size and after.size and np.median(before) < threshold and np.median(after) < threshold:
            return group[1]
    return None


def _fft_correlate(a: np.ndarray, b: np.ndarray) -> np.ndarray:
    size = len(a) + len(b) - 1
    fft_size = 1 << (size - 1).bit_length()
    return np.fft.irfft(np.fft.rfft(a, fft_size) * np.fft.rfft(b[::-1], fft_size), fft_size)[:size]


def _rescale(values: np.ndarray, scale: float) -> np.ndarray:
    size = max(2, int(round(len(values) * scale)))
    return np.interp(
        np.linspace(0, len(values) - 1, size),
        np.arange(len(values)),
        values,
    )


def _overlap(a: np.ndarray, b: np.ndarray, lag: int) -> tuple[np.ndarray, np.ndarray]:
    a_start = max(0, lag)
    b_start = max(0, -lag)
    count = min(len(a) - a_start, len(b) - b_start)
    if count <= 0:
        return np.empty(0), np.empty(0)
    return a[a_start : a_start + count], b[b_start : b_start + count]


def align_envelopes(
    imu: np.ndarray,
    c3d_motion: np.ndarray,
    *,
    imu_cue: int | None = None,
    c3d_cue: int | None = None,
) -> Alignment:
    a = _robust_normalize(c3d_motion)
    best: tuple[float, int, float, np.ndarray] | None = None
    scales = np.linspace(0.9975, 1.0025, 11)
    for scale in scales:
        b = _robust_normalize(_rescale(imu, float(scale)))
        correlation = _fft_correlate(a, b)
        lags = np.arange(-(len(b) - 1), len(a))
        overlap_count = np.minimum(len(a), lags + len(b)) - np.maximum(0, lags)
        valid = overlap_count >= max(100, min(len(a), len(b)) // 3)
        if imu_cue is not None and c3d_cue is not None:
            cue_lag = c3d_cue - int(round(imu_cue * scale))
            valid &= np.abs(lags - cue_lag) <= 25
        score = np.full_like(correlation, -np.inf, dtype=np.float64)
        score[valid] = correlation[valid] / overlap_count[valid]
        index = int(np.argmax(score))
        candidate = float(score[index])
        if best is None or candidate > best[0]:
            best = (candidate, int(lags[index]), float(scale), score)
    if best is None:
        return Alignment(0, 1.0, 0.0, 0.0, 0.0, "low", 999)
    _, lag, scale, scores = best
    scaled = _robust_normalize(_rescale(imu, scale))
    left, right = _overlap(a, scaled, lag)
    coefficient = float(np.corrcoef(left, right)[0, 1]) if len(left) >= 2 else 0.0
    coefficient = coefficient if math.isfinite(coefficient) else 0.0
    finite_scores = np.sort(scores[np.isfinite(scores)])
    margin = float(finite_scores[-1] - finite_scores[-2]) if len(finite_scores) > 1 else 0.0
    # Repeat the lag estimate on the dominant middle half as a stability test.
    q1, q3 = len(left) // 4, 3 * len(left) // 4
    repeatability = 0
    if q3 - q1 >= 100:
        local = _fft_correlate(left[q1:q3], right[q1:q3])
        local_lag = int(np.argmax(local) - (len(right[q1:q3]) - 1))
        repeatability = abs(local_lag)
    if coefficient >= 0.60 and repeatability <= 1:
        confidence = "high"
    elif coefficient >= 0.55 or (coefficient >= 0.30 and repeatability <= 3):
        confidence = "medium"
    else:
        confidence = "low"
    return Alignment(
        lag, scale, (scale - 1.0) * 1_000_000.0,
        coefficient, margin, confidence, repeatability,
    )


def activity_interval(envelope: np.ndarray, rate_hz: int = 100) -> tuple[int, int]:
    if not len(envelope):
        return 0, 0
    median = float(np.median(envelope))
    mad = float(np.median(np.abs(envelope - median)))
    active = envelope > median + max(0.35, 1.5 * 1.4826 * mad)
    active = np.convolve(active.astype(np.int8), np.ones(max(3, rate_hz // 4)), mode="same") > 0
    indices = np.flatnonzero(active)
    if not len(indices):
        return 0, len(envelope) - 1
    pad = rate_hz // 2
    return max(0, int(indices[0]) - pad), min(len(envelope) - 1, int(indices[-1]) + pad)


def interpolate_wheel(wheel: WheelData, grid_us: np.ndarray, max_gap_samples: int = 2) -> tuple[np.ndarray, int]:
    result = np.full((len(grid_us), 6), np.nan, dtype=np.float64)
    long_gaps = 0
    for axis in range(6):
        result[:, axis] = np.interp(grid_us, wheel.time_us, wheel.values[:, axis], left=np.nan, right=np.nan)
    indices = np.searchsorted(wheel.time_us, grid_us, side="right")
    for row, right_index in enumerate(indices):
        if right_index == 0 or right_index >= len(wheel.time_us):
            result[row] = np.nan
            continue
        gap = int(wheel.time_us[right_index] - wheel.time_us[right_index - 1])
        expected = max(1, int(round(wheel.fit_slope_ms * 1000)))
        if gap > expected * (max_gap_samples + 1) + expected // 2:
            result[row] = np.nan
            long_gaps += 1
    return result, long_gaps


def write_csv(path: Path, header: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(header)
        writer.writerows(rows)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def find_c3d_files(folder: Path) -> dict[str, Path]:
    return {path.name.lower(): path for path in folder.rglob("*.c3d")}


def validate_generated_output(output: Path) -> dict[str, object]:
    checked: list[dict[str, object]] = []
    for folder, expected_header in (("raw", RAW_HEADER), ("training", TRAINING_HEADER)):
        for path in sorted((output / folder).glob("*.csv")):
            rows = 0
            previous_time: int | None = None
            with path.open(newline="", encoding="utf-8") as handle:
                reader = csv.reader(handle)
                header = next(reader, None)
                if header != expected_header:
                    raise ValueError(f"Invalid header in {path}: {header}")
                for line_number, row in enumerate(reader, 2):
                    if len(row) != len(expected_header):
                        raise ValueError(f"Invalid column count in {path}:{line_number}")
                    time_us = int(row[0])
                    if previous_time is not None and time_us < previous_time:
                        raise ValueError(f"Non-monotonic time in {path}:{line_number}")
                    if not all(math.isfinite(float(value)) for value in row[1:]):
                        raise ValueError(f"Non-finite value in {path}:{line_number}")
                    previous_time = time_us
                    rows += 1
            checked.append({"file": str(path.relative_to(output)), "rows": rows, "sha256": sha256(path)})
    # The report cannot truthfully include its own SHA-256 because writing the
    # digest changes the file. Validate every generated input to the report and
    # deliberately exclude the report itself from the checked-file manifest.
    validation_report = output / "qc" / "validation_report.json"
    for path in sorted((output / "qc").glob("*.json")):
        if path == validation_report:
            continue
        json.loads(path.read_text(encoding="utf-8"))
        checked.append({"file": str(path.relative_to(output)), "sha256": sha256(path)})
    manifest = output / "dataset_manifest.csv"
    with manifest.open(newline="", encoding="utf-8") as handle:
        manifest_rows = list(csv.DictReader(handle))
    if not manifest_rows:
        raise ValueError("Dataset manifest is empty")
    checked.append({"file": manifest.name, "rows": len(manifest_rows), "sha256": sha256(manifest)})
    report = {"valid": True, "checked_file_count": len(checked), "files": checked}
    validation_report.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return report


def process_trial(xlsx_path: Path, c3d_path: Path | None, output: Path) -> dict[str, object]:
    topic, trial = parse_trial_name(xlsx_path)
    wheels = read_legacy_xlsx(xlsx_path)
    stem = f"{topic}_trial_{trial:02d}_2026-07-10"
    raw_paths: dict[str, str] = {}
    for side, wheel in wheels.items():
        name = f"{stem}_{'left' if side == 'L' else 'right'}_raw.csv"
        path = output / "raw" / name
        write_csv(path, RAW_HEADER, ([int(t), *values] for t, values in zip(wheel.time_us, wheel.values)))
        raw_paths[side] = str(path.relative_to(output))

    row: dict[str, object] = {
        "topic": topic,
        "trial": trial,
        "status": "imu_only" if c3d_path is None else "degraded",
        "ready_to_train": False,
        "imu_source": str(xlsx_path),
        "c3d_source": "" if c3d_path is None else str(c3d_path),
        "raw_left_csv": raw_paths.get("L", ""),
        "raw_right_csv": raw_paths.get("R", ""),
        "training_csv": "",
        "alignment_confidence": "none",
        "correlation": "",
        "offset_frames": "",
        "drift_ppm": "",
        "crop_start_us": "",
        "crop_end_us": "",
        "notes": "missing C3D" if c3d_path is None else "",
    }
    qc: dict[str, object] = {
        "processor_version": PROCESSOR_VERSION,
        "export_schema": 2,
        "topic": topic,
        "trial": trial,
        "imu_sha256": sha256(xlsx_path),
        "ignored_legacy_fields": ["timestamp_utc_ms", "marker"],
        "wheels": {
            side: {
                "samples": int(len(wheel.seq)),
                "sequence_gaps": wheel.gaps,
                "fit_interval_ms": wheel.fit_slope_ms,
                "arrival_fit_rms_ms": wheel.fit_rms_ms,
            }
            for side, wheel in wheels.items()
        },
    }
    if "L" not in wheels or "R" not in wheels:
        row["status"] = "rejected"
        row["notes"] = "missing L or R worksheet"
    else:
        start_us = max(int(wheels["L"].time_us[0]), int(wheels["R"].time_us[0]))
        end_us = min(int(wheels["L"].time_us[-1]), int(wheels["R"].time_us[-1]))
        grid = np.arange(((start_us + 9999) // 10000) * 10000, end_us + 1, 10000, dtype=np.int64)
        left_values, left_long = interpolate_wheel(wheels["L"], grid)
        right_values, right_long = interpolate_wheel(wheels["R"], grid)
        valid = np.isfinite(left_values).all(axis=1) & np.isfinite(right_values).all(axis=1)
        combined_values = (left_values + right_values) / 2.0
        combined_env = imu_envelope(grid[valid], combined_values[valid]) if valid.any() else np.empty(0)
        crop_start, crop_end = 0, max(0, len(grid) - 1)

        if c3d_path is not None and valid.any():
            c3d_env, c3d_rate, c3d_qc = read_c3d_envelope(c3d_path)
            if abs(c3d_rate - 100.0) > 0.01:
                x = np.linspace(0, len(c3d_env) - 1, round(len(c3d_env) * 100.0 / c3d_rate))
                c3d_env = np.interp(x, np.arange(len(c3d_env)), c3d_env)
            imu_cue = detect_three_pulse_cue(combined_env)
            c3d_cue = detect_three_pulse_cue(c3d_env)
            alignment = align_envelopes(combined_env, c3d_env, imu_cue=imu_cue, c3d_cue=c3d_cue)
            imu_active = activity_interval(combined_env)
            c3d_active = activity_interval(c3d_env)
            mapped_start = math.ceil((c3d_active[0] - alignment.lag_frames) / alignment.scale)
            mapped_end = math.floor((c3d_active[1] - alignment.lag_frames) / alignment.scale)
            crop_start = max(imu_active[0], mapped_start, 0)
            crop_end = min(imu_active[1], mapped_end, len(grid) - 1)
            row.update({
                "alignment_confidence": alignment.confidence,
                "correlation": f"{alignment.correlation:.6f}",
                "offset_frames": alignment.lag_frames,
                "drift_ppm": f"{alignment.drift_ppm:.1f}",
                "crop_start_us": int(grid[crop_start]) if crop_start < len(grid) else "",
                "crop_end_us": int(grid[crop_end]) if crop_end < len(grid) else "",
            })
            qc.update({
                "c3d_sha256": sha256(c3d_path),
                "c3d": c3d_qc,
                "alignment": alignment.__dict__,
                "cue": {"imu_frame": imu_cue, "c3d_frame": c3d_cue},
            })
        else:
            alignment = None

        crop_valid = valid.copy()
        crop_valid[: max(0, crop_start)] = False
        if crop_end + 1 < len(crop_valid):
            crop_valid[crop_end + 1 :] = False
        training_path = output / "training" / f"{stem}_training.csv"
        rows = (
            [int(grid[index]), *left_values[index], *right_values[index]]
            for index in np.flatnonzero(crop_valid)
        )
        write_csv(training_path, TRAINING_HEADER, rows)
        row["training_csv"] = str(training_path.relative_to(output))
        long_gaps = left_long + right_long
        if c3d_path is None:
            row["status"] = "imu_only"
        elif alignment is None or alignment.confidence == "low":
            row["status"] = "rejected"
            row["notes"] = "low-confidence C3D alignment"
        elif crop_end <= crop_start:
            row["status"] = "rejected"
            row["notes"] = "no valid shared activity interval"
        elif long_gaps > 0:
            row["status"] = "rejected"
            row["notes"] = f"{long_gaps} grid samples cross gaps longer than two frames"
        elif alignment.confidence == "high":
            row["status"] = "ready"
            row["ready_to_train"] = True
        else:
            row["status"] = "degraded"
            row["notes"] = "medium-confidence C3D alignment"
        qc["training"] = {
            "rows": int(crop_valid.sum()),
            "long_gap_grid_rows": long_gaps,
            "crop_start_index": crop_start,
            "crop_end_index": crop_end,
        }

    qc["result"] = row
    qc_path = output / "qc" / f"{stem}_qc.json"
    qc_path.parent.mkdir(parents=True, exist_ok=True)
    qc_path.write_text(json.dumps(qc, indent=2, ensure_ascii=False), encoding="utf-8")
    return row


def process_dataset(imu_folder: Path, c3d_folder: Path, output: Path) -> list[dict[str, object]]:
    imu_folder = imu_folder.resolve()
    c3d_folder = c3d_folder.resolve()
    output = output.resolve()
    if output == imu_folder or output == c3d_folder or imu_folder in output.parents or c3d_folder in output.parents:
        raise ValueError("Output must be outside both source folders")
    for name in ("raw", "training", "qc"):
        (output / name).mkdir(parents=True, exist_ok=True)
    c3d_files = find_c3d_files(c3d_folder)
    used_c3d: set[Path] = set()
    rows: list[dict[str, object]] = []
    for xlsx in sorted(imu_folder.glob("*.xlsx")):
        topic, trial = parse_trial_name(xlsx)
        expected = C3D_BY_TRIAL.get((topic, trial))
        c3d_path = c3d_files.get(expected.lower()) if expected else None
        if c3d_path is not None:
            used_c3d.add(c3d_path)
        rows.append(process_trial(xlsx, c3d_path, output))
    for c3d_path in sorted(set(c3d_files.values()) - used_c3d):
        rows.append({
            "topic": "", "trial": "", "status": "unmatched", "ready_to_train": False,
            "imu_source": "", "c3d_source": str(c3d_path), "raw_left_csv": "",
            "raw_right_csv": "", "training_csv": "", "alignment_confidence": "none",
            "correlation": "", "offset_frames": "", "drift_ppm": "",
            "crop_start_us": "", "crop_end_us": "", "notes": "unmatched C3D",
        })
    manifest = output / "dataset_manifest.csv"
    fields = list(rows[0].keys()) if rows else []
    write_csv(manifest, fields, ([row.get(field, "") for field in fields] for row in rows))
    summary = {
        "processor_version": PROCESSOR_VERSION,
        "source_imu": str(imu_folder),
        "source_c3d": str(c3d_folder),
        "output": str(output),
        "counts": {
            status: sum(row["status"] == status for row in rows)
            for status in ("ready", "degraded", "imu_only", "unmatched", "rejected")
        },
    }
    (output / "qc" / "processing_summary.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    validate_generated_output(output)
    return rows


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("imu_folder", type=Path)
    parser.add_argument("c3d_folder", type=Path)
    parser.add_argument("output", type=Path)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    arguments = build_parser().parse_args(argv)
    rows = process_dataset(arguments.imu_folder, arguments.c3d_folder, arguments.output)
    counts = {status: sum(row["status"] == status for row in rows) for status in {row["status"] for row in rows}}
    print(json.dumps({"output": str(arguments.output.resolve()), "counts": counts}, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
